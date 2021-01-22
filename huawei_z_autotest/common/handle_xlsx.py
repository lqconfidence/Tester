#！/usr/bin/env python
# -*- coding:utf-8 -*-
# __author__  = admin

import openpyxl,json
import sys
import os
base_path=os.path.dirname(os.getcwd())
sys.path.append(base_path)

class HandleXlsx:

    def __init__(self,file_name):
        self.file_path=base_path + "/data/" + str(file_name)

    def load_execl(self):
        return openpyxl.load_workbook(self.file_path)

    def get_sheet_data(self,index=1):
        sheet_names = self.load_execl().sheetnames
        return self.load_execl()[sheet_names[index]]

    def get_cell_data(self,row,cols):
        return self.get_sheet_data().cell(row=row,column=cols).value

    def get_rows(self):
        return self.get_sheet_data().max_row

    def get_rows_values(self,row):
        row_list=[]
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self,row,cols,value):
        wb=self.load_execl()
        wr=wb.active
        wr.cell(row,cols,value)
        wb.save(self.file_path)

    def get_columns_values(self,col='A'):
        column_list=[]
        for i in self.get_sheet_data()[col]:
            column_list.append(i.value)
        del column_list[0]
        return column_list

    def get_row_numbr_with_caseID(self,case_name):
        '''根据用例编号获取行号'''
        num = 2
        for i in self.get_columns_values():
            if i == case_name:
                return num
            num = num + 1
        if case_name not in self.get_columns_values():
            print("{}不在该列中".format(case_name))

    def get_excel_data(self):
        excel_date=[]
        for i in range(self.get_rows()):
            excel_date.append(self.get_rows_values(i+2))
        return excel_date[:-1]

# handle_excel=HandleXlsx()

if __name__=="__main__":
    handle_excel = HandleXlsx("callback_public.xlsx")
    # data=handle_excel.get_excel_data()
    # data=handle_excel.get_columns_values('M')
    handle_excel.excel_write_data(2, 14, "Fail")

