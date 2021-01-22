import json
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from common.setting import Config


class ExcelHandler():
    def __init__(self, file):
        self.file = file
        self.res = lambda sheet: sheet

    def open_sheet(self, name) -> Worksheet:
        wb = load_workbook(self.file)
        sheet = wb[name]
        wb.close()
        return sheet

    def header(self, sheet_name):
        return [i.value for i in self.open_sheet(sheet_name)[1]]

    def read(self, sheet_name):
        print("开始读取excel")
        data = [dict(zip(self.header(sheet_name),[cell.value for cell in row])) for row in list(self.open_sheet(sheet_name).rows)[1:]]
        # print("===========There is Data====================")
        # <class 'list'>
        # print(type(data))
        return data

    def write(self, sheet_name, row, column, data):
        wb = load_workbook(self.file)
        sheet = wb[sheet_name]
        sheet.cell(row, column).value = data
        wb.save(self.file)
        wb.close()


excel = ExcelHandler(Config.excel_path)

excel_static_privacy = ExcelHandler(Config.excel_static_privacy)
excel_callback_public = ExcelHandler(Config.excel_callback_public)
excel_ppcrawler = ExcelHandler(Config.excel_ppcawler)
excel_download = ExcelHandler(Config.excel_download)
excel_analog = ExcelHandler(Config.excel_analog)
excel_hw_privacyreview_case = ExcelHandler(Config.excel_hw_privacyreview_case)
excel_hw_privacyreview_all = ExcelHandler(Config.excel_hw_privacyreview_all)
excel_app_function = ExcelHandler(Config.excel_function_app)
excel_hw_zhongduan = ExcelHandler(Config.excel_hw_zhongduan)
excel_aosp = ExcelHandler(Config.excel_aosp)
excel_prod = ExcelHandler(Config.excel_prod_public)
#excel_analysis_pa = ExcelHandler(Config.ex)
if __name__ == '__main__':
    # 生成嵌套字典的列表
    data = excel.read('keywordlist')
    print(data)
