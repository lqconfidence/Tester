from openpyxl import load_workbook
from scripts.handle_config import do_config

class HandleExcel:
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_cases(self):
        wb = load_workbook(self.filename)  # 返回一个文件 openpyxl支持几种表格文件
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        head_data_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        one_list = []
        for one_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
            one_list.append(dict(zip(head_data_tuple, one_tuple)))

        return one_list

    def write_result(self, row, actual, result, req_id):
        other_wb = load_workbook(self.filename)
        if self.sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[self.sheetname]   # 获取表格中的sheet
        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=do_config.get_int("excel", "actual_col"), value=actual)  # 在制定行和列中插入内容
            other_ws.cell(row=row, column=do_config.get_int("excel", "result_col"), value=result)
            other_ws.cell(row=row, column=do_config.get_int("excel", "req_id_col"), value=req_id)
            other_wb.save(self.filename)
            other_wb.close()
        else:
            print("传入的行号有误，行号应为大于1的整数")

    def write_final(self, row, score, evaluative):
        other_wb = load_workbook(self.filename)
        if self.sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[self.sheetname]
        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=do_config.get_int("excel", "score_col"), value=score)
            other_ws.cell(row=row, column=do_config.get_int("excel", "evaluative_col"), value=evaluative)
            other_wb.save(self.filename)
            other_wb.close()
        else:
            print("传入的行号有误，行号应为大于1的整数")

    def write_detail(self, row, experience, influence, cert_cycle, cert_random, package_random, risk, virus, score):
        other_wb = load_workbook(self.filename)
        if self.sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[self.sheetname]
        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=do_config.get_int("excel", "experience_col"), value=experience)
            other_ws.cell(row=row, column=do_config.get_int("excel", "influence_col"), value=influence)
            other_ws.cell(row=row, column=do_config.get_int("excel", "cert_cycle_col"), value=cert_cycle)
            other_ws.cell(row=row, column=do_config.get_int("excel", "cert_random_col"), value=cert_random)
            other_ws.cell(row=row, column=do_config.get_int("excel", "package_random_col"), value=package_random)
            other_ws.cell(row=row, column=do_config.get_int("excel", "risk_col"), value=risk)
            other_ws.cell(row=row, column=do_config.get_int("excel", "virus_col"), value=virus)
            other_ws.cell(row=row, column=do_config.get_int("excel", "score_col"), value=score)
            other_wb.save(self.filename)
            other_wb.close()
        else:
            print("传入的行号有误，行号应为大于1的整数")


